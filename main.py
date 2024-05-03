import pandas as pd
import math
from openpyxl import load_workbook
import sys
try:
    delivery_file = pd.ExcelFile('DELIVERY.xlsx')
    delivery_sheet = delivery_file.parse('30')
except FileNotFoundError:
    print("The file name should be 'DELIVERY.xlsx'")
    sys.exit()
except Exception as e:
    print("An error occurred when opening the file 'DELIVERY.xlsx'",str(e))
    sys.exit()
for index, row in delivery_sheet.iterrows():
    po_number=row['PO NO']
    if not math.isnan(po_number):
        art=row['ART']
        clr=row['CLR']
        grids=str(row['GRID'])
        if 'X' in grids:
            grid_split=grids.split('X')
            min_grid=int(grid_split[0])
            max_grid=int(grid_split[1])
            grid=[]
            pos=10
            for i in range(min_grid,max_grid+1):
                left=str(i)+'L'
                right=str(i)+'R'
                left_val=row.iloc[pos]
                right_val=row.iloc[pos+1]
                grid.append((left,left_val))
                grid.append((right,right_val))
                pos+=2
        else:
            print("GRID TYPO ERROR!!!")
        if art and clr and min_grid and max_grid:
            try:
                stiching_file = pd.ExcelFile('STICHING_TODAY.xlsx')
                stiching_sheet = stiching_file.parse('PENDING PLAN',header=0)
                book = load_workbook('STICHING_TODAY.xlsx')
                sheet = book.active
            except FileNotFoundError:
                print("The file name should be 'STICHING_TODAY.xlsx'")
                sys.exit()
            except Exception as e:
                print(
                    "An error occurred when opening the file 'STICHING_TODAY.xlsx'",str(e))
                sys.exit()
            for st_index, st_row in stiching_sheet.iterrows():
                if 'Pur. Doc.' in stiching_sheet.columns:
                    today_quantity=st_row['Quantity']
                    today_po=st_row['Pur. Doc.']
                    today_supplier=st_row['Supplier/Supplying Plant']
                    today_short_text=st_row['Short Text']
                    if today_po == po_number:
                        if art in st_row['Short Text'] and clr in st_row['Material']:
                            for tup in grid:
                                try:
                                    variant, val = tup
                                except ValueError:
                                    print(f"Error unpadking tuple:{tup} {grid}")
                                if variant in st_row['Short Text'] and val == st_row['Quantity']:
                                    print("WE FOUND EXACT MATCH HERE!!!!")
                                    sheet.cell(row=st_index + 2, column=stiching_sheet.columns.get_loc('Remark') + 1,
                                               value="TODAY DELIVERY")
                    print("Remark of rows:", st_row['Remark'])
                    if st_row['Remark'] != 'TODAY DELIVERY':
                        try:
                            stiching_file_yesterday = pd.ExcelFile('STICHING_YESTERDAY.xlsx')
                            stiching_sheet_yesterday = stiching_file_yesterday.parse('PENDING PLAN',header=0)
                        except FileNotFoundError:
                            print(
                                "The file name of yesterdays should be 'STICHING_YESTERDAY.xlsx'")
                            sys.exit()
                        except Exception as e:
                            print(
                                "An error occurred when opening the file 'STICHING_YESTERDAY.xlsx'",
                                str(e))
                            sys.exit()
                        for index, row in stiching_sheet_yesterday.iterrows():
                            yesterday_po = row['Pur. Doc.']
                            yesterday_supplier = row['Supplier/Supplying Plant']
                            yesterday_short_text = row['Short Text']
                            yesterday_quantity = row['Quantity']
                            yesterday_remark = row['Remark']
                            if yesterday_remark == 'COUNTING BALANCE' or yesterday_remark == 'TODAY DELIVERY':
                                if  today_quantity == yesterday_quantity and today_po == yesterday_po and today_supplier == yesterday_supplier and today_short_text == yesterday_short_text:
                                    sheet.cell(row=st_index + 2,
                                               column=stiching_sheet.columns.get_loc('Remark') + 1,
                                               value="COUNTING BALANCE")
                else:
                    print("'Pur. Doc.' does not exist in the STICHING_TODAY \nThe header of sheet should be the first line")
                    print("PRINTING COLUMNS:", stiching_sheet.columns)
            book.save('STICHING_TODAY.xlsx')
