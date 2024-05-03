import pandas as pd
import math
from openpyxl import load_workbook
import sys

try:
    delivery_file = pd.ExcelFile('DELIVERY.xlsx')
    delivery_sheet = delivery_file.parse('02')
except FileNotFoundError:
    print("The file name should be 'DELIVERY.xlsx'")
    sys.exit()
except Exception as e:
    print("An error occurred when opening the file 'DELIVERY.xlsx'\n Check the sheet name. It should be '02'", str(e))
    sys.exit()

try:
    stiching_file_today = pd.ExcelFile('STITCHING_TODAY.xlsx')
    stiching_sheet_today = stiching_file_today.parse('PENDING PLAN', header=0)
    stiching_sheet_today['Excel Row'] = range(2, len(stiching_sheet_today) + 2)
except FileNotFoundError:
    print("The file name of today's stiching report should be 'STITCHING_TODAY.xlsx'")
    sys.exit()
except Exception as e:
    print("An error occurred when opening the file 'STITCHING_TODAY.xlsx'\n Check the sheet name. It should be 'PENDING PLAN'", str(e))
    sys.exit()

try:
    stiching_file_yesterday = pd.ExcelFile('STITCHING_YESTERDAY.xlsx')
    stiching_sheet_yesterday = stiching_file_yesterday.parse('PENDING PLAN', header=0)
except FileNotFoundError:
    print("The file name of yesterday's stitching report should be 'STITCHING_YESTERDAY.xlsx'")
    sys.exit()
except Exception as e:
    print("An error occurred when opening the file 'STITCHING_YESTERDAY.xlsx'\n Check the sheet name. It should be 'PENDING PLAN'", str(e))
    sys.exit()
# Group by 'Pur. Doc.'
try:
    stiching_today_groups = stiching_sheet_today.groupby('Pur. Doc.')
    stiching_yesterday_groups = stiching_sheet_yesterday.groupby('Pur. Doc.')
except KeyError:
    print("\nThe column 'Pur. Doc.' does not exist in stitching file. Please check the column names.\n*The header should start from 1st row.*")
    sys.exit()
# Load the workbook with openpyxl
book = load_workbook('STITCHING_TODAY.xlsx')
sheet = book.active

# Process each delivery row
for index, delivery_row in delivery_sheet.iterrows():
    po_number = delivery_row['PO NO']
    if isinstance(po_number,(int,float)) and not math.isnan(po_number):
        art = delivery_row['ART']
        clr = delivery_row['CLR']
        grids = str(delivery_row['GRID'])
        if 'X' in grids:
            grid_split = grids.split('X')
            min_grid = int(grid_split[0])
            max_grid = int(grid_split[1])
            grid = []
            pos = 10
            for i in range(min_grid, max_grid + 1):
                left = str(i) + 'L'
                right = str(i) + 'R'
                left_val = delivery_row.iloc[pos]
                right_val = delivery_row.iloc[pos + 1]
                grid.append((left, left_val))
                grid.append((right, right_val))
                pos += 2
        else:
            print("GRID TYPO ERROR!!!")
   #--------------------------------------





        # Process today's stitching rows
        if po_number in stiching_today_groups.groups and art and clr and min_grid and max_grid:
            for st_index, st_row in stiching_today_groups.get_group(po_number).iterrows():
                if 'Pur. Doc.' in stiching_today_groups.get_group(po_number).columns:
                    today_quantity = st_row['Quantity']
                    today_po = st_row['Pur. Doc.']
                    today_supplier = st_row['Supplier/Supplying Plant']
                    today_short_text = st_row['Short Text']
                    if today_po == po_number:
                        if art in st_row['Short Text'] and clr in st_row['Material']:
                            for tup in grid:
                                try:
                                    variant, val = tup
                                except ValueError:
                                    print(f"Error unpacking tuple:{tup} {grid}")
                                if variant in st_row['Short Text'] and val == st_row['Quantity']:
                                    # print("WE FOUND EXACT MATCH HERE!!!!")
                                    excel_row = st_row['Excel Row']
                                    sheet.cell(row=excel_row,
                                               column=stiching_today_groups.get_group(po_number).columns.get_loc(
                                                   'Remark') + 1,
                                               value="TODAY DELIVERY")

                # Process yesterday's stitching rows
                # if po_number in stiching_yesterday_groups.groups:
                #     for st_index_yesterday, st_row_yesterday in stiching_yesterday_groups.get_group(po_number).iterrows():
                #         yesterday_remark=st_row_yesterday['Remark']
                #         if yesterday_remark=='TODAY DELIVERY' or yesterday_remark=='COUNTING BALANCE':
                #             yesterday_quantity = st_row_yesterday['Quantity']
                #             yesterday_supplier = st_row_yesterday['Supplier/Supplying Plant']
                #             yesterday_short_text = st_row_yesterday['Short Text']
                #
                #             if yesterday_quantity==today_quantity and yesterday_supplier==today_supplier and yesterday_short_text==today_short_text:
                #                 print("WE FOUND COUNTING BALANCE HERE!!!!")
                #                 excel_row = st_row['Excel Row']
                #                 sheet.cell(row=excel_row,
                #                            column=stiching_today_groups.get_group(po_number).columns.get_loc(
                #                                'Remark') + 1,
                #                            value="COUNTING BALANCE")

for index_today,row in stiching_sheet_today.iterrows():
    po_number_today=row['Pur. Doc.']
    today_quantity = row['Quantity']
    today_po = row['Pur. Doc.']
    today_supplier = row['Supplier/Supplying Plant']
    today_short_text = row['Short Text']
    if index_today>50:
        print(index_today,"Today")
    if row['Remark']!="TODAY DELIVERY" and not math.isnan(po_number_today) and po_number_today in stiching_yesterday_groups.groups:
        for st_index_yesterday, st_row_yesterday in stiching_yesterday_groups.get_group(po_number_today).iterrows():
            yesterday_remark = st_row_yesterday['Remark']
            if st_index_yesterday>50 and index<100:
                    print(index,"Yesterday")

            if yesterday_remark == 'TODAY DELIVERY' or yesterday_remark == 'COUNTING BALANCE':
                print(yesterday_remark)
                yesterday_quantity = st_row_yesterday['Quantity']
                yesterday_supplier = st_row_yesterday['Supplier/Supplying Plant']
                yesterday_short_text = st_row_yesterday['Short Text']



                if yesterday_quantity == today_quantity and yesterday_supplier == today_supplier and yesterday_short_text == today_short_text:
                    print("WE FOUND COUNTING BALANCE HERE!!!!")
                    excel_row = row['Excel Row']
                    sheet.cell(row=excel_row,
                               column=stiching_today_groups.get_group(po_number_today).columns.get_loc(
                                   'Remark') + 1,
                               value="COUNTING BALANCE")


try:

    # Save the changes
    book.save('STITCHING_TODAY.xlsx')
except PermissionError:
    print("STITCHING TODAY file has permission error.\nPlease close the file STITCHING_TODAY.xlsx if it is already opened and try again.\n\n!!PROCESS NOT COMPLETED!!")